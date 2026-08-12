import os
import smtplib
import logging
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from pathlib import Path
from django.conf import settings
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from icalendar import Calendar, Event, vCalAddress, vText
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


class EmailService:
    """Handle email notifications."""

    @staticmethod
    def send_appointment_confirmation(appointment):
        """Send appointment confirmation email to patient."""
        patient = appointment.patient
        doctor = appointment.doctor

        subject = f"Appointment Confirmation - Dr. {doctor.name}"

        context = {
            'patient_name': patient.name,
            'doctor_name': doctor.name,
            'doctor_specialization': doctor.specialization,
            'date': appointment.appointment_date.strftime('%d/%m/%Y'),
            'time': appointment.appointment_time.strftime('%I:%M %p'),
            'problem': appointment.problem_description,
            'appointment_id': appointment.id,
            'status': appointment.get_status_display(),
        }

        try:
            html_message = render_to_string('appointments/email/confirmation.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject,
                plain_message,
                settings.ADMIN_EMAIL,
                [patient.email],
                html_message=html_message,
                fail_silently=False,
            )
            logger.info(f"Email sent to {patient.email} for appointment {appointment.id}")
            return True
        except Exception as e:
            logger.error(f"Failed to send email: {str(e)}")
            return False

    @staticmethod
    def send_doctor_notification(appointment):
        """Send notification to doctor about new appointment."""
        patient = appointment.patient
        doctor = appointment.doctor

        subject = f"New Appointment - {patient.name}"

        context = {
            'patient_name': patient.name,
            'patient_age': patient.age,
            'patient_gender': patient.get_gender_display(),
            'patient_phone': patient.phone,
            'patient_email': patient.email,
            'doctor_name': doctor.name,
            'date': appointment.appointment_date.strftime('%d/%m/%Y'),
            'time': appointment.appointment_time.strftime('%I:%M %p'),
            'problem': appointment.problem_description,
            'appointment_id': appointment.id,
        }

        try:
            html_message = render_to_string('appointments/email/doctor_notification.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject,
                plain_message,
                settings.ADMIN_EMAIL,
                [doctor.email],
                html_message=html_message,
                fail_silently=False,
            )
            logger.info(f"Doctor notification sent to {doctor.email}")
            return True
        except Exception as e:
            logger.error(f"Failed to send doctor notification: {str(e)}")
            return False

    @staticmethod
    def send_cancellation_email(appointment):
        """Notify patient that their appointment was cancelled."""
        patient = appointment.patient
        doctor = appointment.doctor

        subject = f"Appointment Cancelled - Dr. {doctor.name}"

        context = {
            'patient_name': patient.name,
            'doctor_name': doctor.name,
            'doctor_specialization': doctor.specialization,
            'date': appointment.appointment_date.strftime('%d/%m/%Y'),
            'time': appointment.appointment_time.strftime('%I:%M %p'),
            'appointment_id': appointment.id,
        }

        try:
            html_message = render_to_string('appointments/email/cancellation.html', context)
            plain_message = strip_tags(html_message)

            send_mail(
                subject,
                plain_message,
                settings.ADMIN_EMAIL,
                [patient.email],
                html_message=html_message,
                fail_silently=False,
            )
            logger.info(f"Cancellation email sent to {patient.email} for appointment {appointment.id}")
            return True
        except Exception as e:
            logger.error(f"Failed to send cancellation email: {str(e)}")
            return False


class SMSService:
    """Handle SMS notifications."""

    @staticmethod
    def send_appointment_confirmation(appointment):
        """Send SMS confirmation to patient."""
        patient = appointment.patient
        doctor = appointment.doctor

        message = f"""
        Medical: Appointment Confirmed!
        Patient: {patient.name}
        Doctor: Dr. {doctor.name} ({doctor.specialization})
        Date: {appointment.appointment_date.strftime('%d/%m/%Y')}
        Time: {appointment.appointment_time.strftime('%I:%M %p')}
        Thank you for choosing our service.
        """

        if settings.SMS_MOCK:
            logger.info(f"[MOCK SMS] To: {patient.phone} | Message: {message}")
            print(f"\n[=== SMS MOCK ===]")
            print(f"To: {patient.phone}")
            print(f"Message: {message}")
            print(f"[=== END SMS ===]\n")
            return True
        else:
            try:
                from twilio.rest import Client

                phone = patient.phone.strip()
                if not phone.startswith('+'):
                    phone = '+91' + phone.lstrip('0')

                client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)
                client.messages.create(
                    body=message,
                    from_=settings.TWILIO_PHONE_NUMBER,
                    to=phone
                )
                logger.info(f"SMS sent to {phone}")
                return True
            except Exception as e:
                logger.error(f"Failed to send SMS: {str(e)}")
                return False


class CalendarService:
    """Handle calendar integration (ICS generation)."""

    @staticmethod
    def generate_ics(appointment):
        """Generate ICS file for appointment."""
        cal = Calendar()
        cal.add('prodid', '-//Medical Appointment//mxm.dk//')
        cal.add('version', '2.0')

        event = Event()
        event.add('summary', f"Appointment: {appointment.patient.name} - Dr. {appointment.doctor.name}")
        event.add('description', f"""
        Patient: {appointment.patient.name}
        Age: {appointment.patient.age}
        Gender: {appointment.patient.get_gender_display()}
        Problem: {appointment.problem_description}
        Contact: {appointment.patient.phone} | {appointment.patient.email}
        """)

        start_datetime = datetime.combine(appointment.appointment_date, appointment.appointment_time)
        end_datetime = start_datetime + timedelta(hours=1)

        event.add('dtstart', start_datetime)
        event.add('dtend', end_datetime)
        event.add('dtstamp', datetime.now())

        event.add('location', 'Medical Clinic')

        organizer = vCalAddress(f'MAILTO:{settings.ADMIN_EMAIL}')
        organizer.params['cn'] = vText('Medical Appointment')
        event.add('organizer', organizer)

        cal.add_component(event)

        return cal.to_ical().decode('utf-8')

    @staticmethod
    def save_ics(appointment):
        """Save ICS file to disk and return path."""
        ics_content = CalendarService.generate_ics(appointment)
        filename = f"appointment_{appointment.id}_{appointment.appointment_date.strftime('%Y%m%d')}.ics"
        filepath = Path(settings.MEDIA_ROOT) / 'calendar' / filename

        os.makedirs(filepath.parent, exist_ok=True)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(ics_content)

        return filepath


class ExcelService:
    """Generate Excel reports for doctors."""

    @staticmethod
    def generate_doctor_report(doctor, appointments=None):
        """Generate Excel report for a doctor."""
        if appointments is None:
            appointments = doctor.appointments.filter(status__in=['confirmed', 'completed'])

        wb = Workbook()
        ws = wb.active
        ws.title = "Appointments"

        headers = [
            'Sl No', 'Patient Name', 'Age', 'Gender', 'Phone', 'Email',
            'Appointment Date', 'Appointment Time', 'Problem', 'Status', 'Booked Via'
        ]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0A4B6E', end_color='0A4B6E', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for idx, appointment in enumerate(appointments, 1):
            row = idx + 1
            patient = appointment.patient

            data = [
                idx,
                patient.name,
                patient.age,
                patient.get_gender_display(),
                patient.phone,
                patient.email,
                appointment.appointment_date.strftime('%d/%m/%Y'),
                appointment.appointment_time.strftime('%I:%M %p'),
                appointment.problem_description[:50] + ('...' if len(appointment.problem_description) > 50 else ''),
                appointment.get_status_display(),
                'Voice' if appointment.voice_booking else 'Manual'
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        for col in range(1, len(headers) + 1):
            max_length = 0
            column = ws.column_dimensions[chr(64 + col)]
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            column.width = max_length + 5

        ws_summary = wb.create_sheet("Summary")
        ws_summary.cell(row=1, column=1, value="Doctor Name")
        ws_summary.cell(row=1, column=2, value=doctor.name)
        ws_summary.cell(row=2, column=1, value="Specialization")
        ws_summary.cell(row=2, column=2, value=doctor.specialization)
        ws_summary.cell(row=3, column=1, value="Total Appointments")
        ws_summary.cell(row=3, column=2, value=len(appointments))
        ws_summary.cell(row=4, column=1, value="Generated On")
        ws_summary.cell(row=4, column=2, value=datetime.now().strftime('%d/%m/%Y %H:%M'))

        filename = f"doctor_{doctor.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = Path(settings.MEDIA_ROOT) / 'excel' / filename
        os.makedirs(filepath.parent, exist_ok=True)
        wb.save(filepath)

        return filepath, filename

    @staticmethod
    def generate_all_doctors_report():
        """Generate a combined report for all doctors."""
        from .models import Doctor

        wb = Workbook()
        ws = wb.active
        ws.title = "All Appointments"

        headers = [
            'Sl No', 'Doctor', 'Specialization', 'Patient Name', 'Age', 'Gender',
            'Phone', 'Email', 'Date', 'Time', 'Problem', 'Status', 'Booked Via'
        ]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0A4B6E', end_color='0A4B6E', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        all_appointments = []
        for doctor in Doctor.objects.filter(is_active=True):
            appointments = doctor.appointments.filter(status__in=['confirmed', 'completed'])
            for appt in appointments:
                all_appointments.append({
                    'doctor': doctor,
                    'appointment': appt
                })

        for idx, item in enumerate(all_appointments, 1):
            row = idx + 1
            doctor = item['doctor']
            appointment = item['appointment']
            patient = appointment.patient

            data = [
                idx,
                doctor.name,
                doctor.specialization,
                patient.name,
                patient.age,
                patient.get_gender_display(),
                patient.phone,
                patient.email,
                appointment.appointment_date.strftime('%d/%m/%Y'),
                appointment.appointment_time.strftime('%I:%M %p'),
                appointment.problem_description[:40] + ('...' if len(appointment.problem_description) > 40 else ''),
                appointment.get_status_display(),
                'Voice' if appointment.voice_booking else 'Manual'
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        for col in range(1, len(headers) + 1):
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[chr(64 + col)].width = max_length + 5

        filename = f"all_doctors_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = Path(settings.MEDIA_ROOT) / 'excel' / filename
        os.makedirs(filepath.parent, exist_ok=True)
        wb.save(filepath)

        return filepath, filename


class TTSService:
    """Generate natural-sounding speech using Google Cloud TTS."""

    _client = None

    @classmethod
    def get_client(cls):
        if cls._client is None:
            from google.cloud import texttospeech
            cls._client = texttospeech.TextToSpeechClient()
        return cls._client

    @staticmethod
    def synthesize_speech(text, voice_name='en-US-Neural2-F'):
        """
        Convert text to speech audio (MP3 bytes) using Google Cloud TTS.
        voice_name options include:
          en-US-Neural2-F (female), en-US-Neural2-D (male),
          en-US-Wavenet-F, en-US-Wavenet-D, etc.
        """
        from google.cloud import texttospeech

        client = TTSService.get_client()

        synthesis_input = texttospeech.SynthesisInput(text=text)

        voice = texttospeech.VoiceSelectionParams(
            language_code='en-US',
            name=voice_name,
        )

        audio_config = texttospeech.AudioConfig(
            audio_encoding=texttospeech.AudioEncoding.MP3,
            speaking_rate=1.0,
            pitch=0.0,
        )

        response = client.synthesize_speech(
            input=synthesis_input,
            voice=voice,
            audio_config=audio_config,
        )

        return response.audio_content